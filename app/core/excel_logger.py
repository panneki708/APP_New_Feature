import os
import stat
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.core.logger import logger, log_function


class ExcelLogger:
    def __init__(self, file_path=os.path.join('C:/tmp', f'test_station_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')):
        self.file_path = file_path
        self.workbook = None
        self.unit_sheet = None
        self.interlock_sheet = None
        self.self_test_sheet = None
        self.BNC_sheet = None
        self.resistance_sheet = None
        self.Imp_sheet = None
        self.logger1 = logger.getChild('ExcelLogger')
        self.pn=''
        self.sn=''
        self.excel_time=datetime.now().strftime("%Y%m%d_%H%M%S")

        # Create the directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        # Initialize or load the workbook
        self._init_workbook()

    @log_function
    def _init_workbook(self):
        """Initialize or load the workbook"""
        if os.path.exists(self.file_path):
            try:
                self.workbook = load_workbook(self.file_path)
            except Exception as e:
                self.logger1.error(f"Error loading existing workbook:", exc_info=True,
                                   extra={'func_name': 'init_workbook'})
                self._create_new_workbook()
        else:
            self._create_new_workbook()

    @log_function
    def _create_new_workbook(self):
        """Create a new empty workbook"""
        self.workbook = Workbook()

        # Instead of removing the default sheet, just rename it to one of your expected sheets
        default_sheet = self.workbook.active
        default_sheet.title = "Unit Setup"  # Or whichever sheet you expect to use first

        # Create headers for the default sheet
        self._create_unit_headers(default_sheet)

        try:
            self.workbook.save(self.file_path)
            self.logger1.info(f"Created new Excel file at {self.file_path}",
                              extra={'func_name': 'create_new_workbook'})
        except Exception as e:
            self.logger1.error(f"Error saving new workbook: {e}", exc_info=True,
                               extra={'func_name': 'create_new_workbook'})

    @staticmethod
    def _freq_to_sheet_suffix(freq_text):
        """Convert frequency text to a safe sheet name suffix e.g. '60 MHz' -> '60MHz'"""
        return freq_text.replace(' ', '').replace('.', '_')

    @log_function
    def reset_sheet(self, sheet_name):
        """Clear all data from a specific sheet (except headers) and recreate headers if needed"""
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.logger1.warning(f"Sheet '{sheet_name}' does not exist",
                                     extra={'func_name': 'reset_sheet'})
                return False

            sheet = self.workbook[sheet_name]

            # "Unit Setup" has no header row, so clear from row 1;
            # all other sheets keep the header in row 1.
            if sheet_name == "Unit Setup":
                if sheet.max_row >= 1:
                    sheet.delete_rows(1, sheet.max_row)
            elif sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row)  # Delete from row 2 to end

            # Reapply headers based on sheet type
            if sheet_name == "Interlock Test":
                self._create_interlock_headers(sheet)
            elif sheet_name == "Self Test":
                self._create_self_test_headers(sheet)
            elif sheet_name == "Zone1-Inner_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone2-Mid_Inner_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone3-Mid_Edge_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone4-Edge_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone5-Outer_Res_scan":
                self._create_resistance_headers(sheet)
            elif "_Imp_scan" in sheet_name:
                self._create_impedance_headers(sheet)
            elif sheet_name == "BNC Port Verification":
                self._create_BNC_headers(sheet)
            elif sheet_name == "Unit Setup":
                self._create_unit_headers(sheet)

            self.workbook.save(self.file_path)
            self.logger1.info(f"Reset sheet '{sheet_name}' successfully",
                              extra={'func_name': 'reset_sheet'})
            return True
        except Exception as e:
            self.logger1.error(f"Error resetting sheet '{sheet_name}': {e}", exc_info=True,
                               extra={'func_name': 'reset_sheet'})
            return False

    def _ensure_sheet_exists(self, sheet_name, create_headers_func):
        """Ensure a sheet exists, creating it if necessary"""
        if sheet_name not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(sheet_name)
            create_headers_func(sheet)
            self.workbook.save(self.file_path)
            return sheet
        return self.workbook[sheet_name]

    def _create_unit_headers(self, sheet):
        """Create headers for the unit setup sheet"""
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30

    def _create_interlock_headers(self, sheet):
        """Create headers for the interlock test sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Test Name
        sheet.column_dimensions['C'].width = 15  # Open Count
        sheet.column_dimensions['D'].width = 10  # Closed Count
        sheet.column_dimensions['E'].width = 15  # Test Result
        sheet.column_dimensions['F'].width = 30  # Notes

        headers = [
            "Timestamp", "Test Name", "Open Count", "Closed Count",
            "Test Result", "Notes"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_self_test_headers(self, sheet):
        """Create headers for the self test sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Unit Identifier
        sheet.column_dimensions['C'].width = 15  # Test Result
        sheet.column_dimensions['D'].width = 30  # Test Details
        sheet.column_dimensions['E'].width = 30  # Notes

        headers = [
            "Timestamp", "Unit Identifier", "Test Result",
            "Test Details", "Notes"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_resistance_headers(self, sheet):
        """Create headers for the resistance measurements sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Zone
        sheet.column_dimensions['C'].width = 15  # Setpoint
        sheet.column_dimensions['D'].width = 15  # Resistance
        sheet.column_dimensions['E'].width = 10  # Status
        sheet.column_dimensions['F'].width = 10  # Table Row

        headers = [
            "Timestamp", "Zone", "Setpoint", "Resistance (Ω)",
            "Status", "Table Row"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_impedance_headers(self, sheet):
        """Create headers for the impedance measurements sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Zone
        sheet.column_dimensions['C'].width = 20  # Frequency
        sheet.column_dimensions['D'].width = 15  # Setpoint
        sheet.column_dimensions['E'].width = 15  # Real
        sheet.column_dimensions['F'].width = 15  # Image
        sheet.column_dimensions['G'].width = 15  # Impedance
        sheet.column_dimensions['H'].width = 10  # Status

        headers = [
            "Timestamp", "Zone", "Frequency", "Setpoint",
            "Real(Ω)", "Imaginary", "Impedance(Z)", "Status"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_BNC_headers(self, sheet):
        """Create headers for the BNC Port Verification sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Zone NAME
        sheet.column_dimensions['C'].width = 20  # TEST_VALUE
        sheet.column_dimensions['D'].width = 15  # STATUS

        headers = [
            "Timestamp", "Zone", "Value(db)", "Status"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_summary_headers(self, sheet):
            """Create headers for the Summary sheet with additional metadata fields"""
            # Set up metadata fields in rows 1-9
            metadata_fields = [
                ("EID", ""),
                ("SERIAL NUMBER", ""),
                ("MODEL NUMBER", ""),
                ("VERSION", ""),
                ("TESTER NAME", ""),
                ("COMMENT", ""),
                ("START TIME", ""),
                ("END TIME", ""),
                ("OVERALL RESULT", ""),
                ("TEST FIXTURE SN", ""),
                ("VNA SN", ""),
                ("ECAL SN", "")
            ]

            # Write metadata fields with formatting
            for row_num, (field, _) in enumerate(metadata_fields, start=1):
                # Field name cell
                sheet.cell(row=row_num, column=1, value=field)
                sheet.cell(row=row_num, column=1).font = Font(bold=True)

                # Value cell (empty initially)
                sheet.cell(row=row_num, column=2, value="")

                # Format for OVERALL RESULT
                if field == "OVERALL RESULT":
                    sheet.cell(row=row_num, column=2).font = Font(bold=True)
                    sheet.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')

            # Add space between metadata and TESTSTEP section
            sheet.row_dimensions[13].height = 15

            # TESTSTEP section (now starting at row 14)
            sheet.cell(row=14, column=1, value="TESTSTEP").font = Font(bold=True, color="FFFFFF")
            sheet.cell(row=14, column=1).fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid")

            sheet.cell(row=14, column=2, value="STATUS").font = Font(bold=True, color="FFFFFF")
            sheet.cell(row=14, column=2).fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid")

            # Add space between sections
            sheet.row_dimensions[21].height = 15

            # STEP section headers (now starting at row 22 from column A)
            headers_row22 = ["Step", "Unit", "Low_Limit", "Measure", "High_Limit",
                             "TestStep", "TestPoints", "Status"]
            for col_num, header in enumerate(headers_row22, start=1):  # Starting at column A (1)
                cell = sheet.cell(row=22, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Set column widths
            sheet.column_dimensions['A'].width = 20  # Step
            sheet.column_dimensions['B'].width = 15  # Unit
            sheet.column_dimensions['C'].width = 15  # Low_Limit
            sheet.column_dimensions['D'].width = 15  # Measure
            sheet.column_dimensions['E'].width = 15  # High_Limit
            sheet.column_dimensions['F'].width = 15  # TestStep
            sheet.column_dimensions['G'].width = 15  # TestPoints
            sheet.column_dimensions['H'].width = 15  # Status

            sheet.freeze_panes = "A23"  # Freeze above the STEP section

    @log_function
    def update_overall_result(self, result, PN='NA', SN='NA'):
        """Update the overall result and rename file accordingly"""
        try:
            result = result.upper()
            if result not in ['PASS', 'FAIL']:
                self.logger1.warning(f"Invalid result: {result}. Must be 'PASS' or 'FAIL'")
                return False

            if PN != 'NA' and SN != 'NA':
                self.pn = PN
                self.sn = SN

            # Create new filename based on result
            new_filename = f"{self.pn}_{self.sn}_{self.excel_time}_{result}.xlsx"
            new_file_path = os.path.join("C:\\tmp", new_filename)
            print("new_fil", new_file_path)

            # If file already exists with different name, rename it
            if self.file_path != new_file_path:
                if os.path.exists(self.file_path):
                    # Close the workbook before renaming
                    self.workbook.close()

                    # Rename the file
                    os.rename(self.file_path, new_file_path)
                    self.file_path = new_file_path

                    # Reopen the workbook
                    self.workbook = load_workbook(self.file_path)
                    self.logger1.info(f"Renamed file to: {new_file_path}",
                                      extra={'func_name': 'update_overall_result'})

            # Make the result file read-only so it cannot be edited after the test run
            if os.path.exists(self.file_path):
                os.chmod(self.file_path,
                         stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
                self.logger1.info(f"Set file as read-only: {self.file_path}",
                                  extra={'func_name': 'update_overall_result'})

            return True

        except Exception as e:
            self.logger1.error(f"Error updating overall result: {e}", exc_info=True,
                               extra={'func_name': 'update_overall_result'})
            return False

    @log_function
    def log_unit_setup(self, unit_data):
        """Log unit setup data to the Excel file"""
        try:
            # Ensure sheet exists
            self.unit_sheet = self._ensure_sheet_exists(
                "Unit Setup",
                self._create_unit_headers
            )

            # Get the next available row
            row_num = 1 if self.unit_sheet.max_row == 1 and all(
                cell.value is None for cell in self.unit_sheet[1]) else self.unit_sheet.max_row + 1

            fields = [
                ("Vendor Name", unit_data.get('Vendor_name', '')),
                ("Fixture Number", unit_data.get('Fixture_number', '')),
                ("Test Operator Name", unit_data.get('test_operator_name', '')),
                ("Test Date", unit_data.get('test_date', '')),
                ("VNA Calibration Date", unit_data.get('vna_calibration_date', '')),
                ("VNA SN", unit_data.get('vna_sn', '')),
                ("Ecal SN", unit_data.get('ecal_sn', '')),
                ("PCB Control Part Number", unit_data.get('pcb_part_number', '')),
                ("PCB Control Revision", unit_data.get('pcb_revision', '')),
                ("PCB Control Serial Number", unit_data.get('pcb_serial_number', '')),
                ("Assembly Part Number", unit_data.get('assembly_part_number', '')),
                ("Assembly Revision", unit_data.get('assembly_revision', '')),
                ("Assembly Serial Number", unit_data.get('assembly_serial_number', '')),
                ("Product ID", unit_data.get('product_id', '')),
                ("ESI Revision", unit_data.get('esi_revision', '')),
                ("Configuration ID", unit_data.get('configuration_id', '')),
                ("EtherCAT Address", unit_data.get('ethercat_address', '')),
                ("Firmware Version", unit_data.get('firmware_version', ''))
            ]

            for i, (field_name, field_value) in enumerate(fields, start=0):
                field_row = row_num + i
                self.unit_sheet.cell(row=field_row, column=1, value=field_name)
                self.unit_sheet.cell(row=field_row, column=1).font = Font(bold=True)
                self.unit_sheet.cell(row=field_row, column=2, value=field_value)
                for col in [1, 2]:
                    self.unit_sheet.cell(row=field_row, column=col).alignment = Alignment(
                        horizontal='left', vertical='center'
                    )

            self.unit_sheet.append([])
            self.workbook.save(self.file_path)
            self.logger1.info(f"Logged unit setup data to {self.file_path}",
                              extra={'func_name': 'log_unit_setup'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging unit setup data: {e}", exc_info=True,
                               extra={'func_name': 'log_unit_setup'})
            return False

    @log_function
    def log_interlock_test(self, test_name, test_passed, open_count, closed_count, notes=""):
        """Log interlock test results to the Excel file"""
        try:
            # Ensure sheet exists
            self.interlock_sheet = self._ensure_sheet_exists(
                "Interlock Test",
                self._create_interlock_headers
            )

            # Get the next available row
            row_num = self.interlock_sheet.max_row + 1

            # Write data with timestamp
            self.interlock_sheet.cell(
                row=row_num, column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            self.interlock_sheet.cell(row=row_num, column=2, value=test_name)

            # Test Result with color coding
            result_cell = self.interlock_sheet.cell(
                row=row_num, column=5,
                value="PASS" if test_passed else "FAIL"
            )
            result_cell.font = Font(bold=True)
            result_cell.fill = PatternFill(
                start_color="00AA00" if test_passed else "FF0000",
                end_color="00AA00" if test_passed else "FF0000",
                fill_type="solid"
            )
            result_cell.alignment = Alignment(horizontal='center')

            self.interlock_sheet.cell(row=row_num, column=3, value=open_count)
            self.interlock_sheet.cell(row=row_num, column=4, value=closed_count)
            self.interlock_sheet.cell(row=row_num, column=6, value=notes)

            # Save the workbook
            self.workbook.save(self.file_path)
            self.logger1.info(f"Logged interlock test result to {self.file_path}",
                              extra={'func_name': 'Log_interlock_test'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging interlock test result: {e}", exc_info=True,
                               extra={'func_name': 'Log_interlock_test'})
            return False

    @log_function
    def log_self_test(self, unit_identifier, test_passed, test_details="", notes=""):
        """Log self test results to the Excel file"""
        try:
            # Ensure sheet exists
            self.self_test_sheet = self._ensure_sheet_exists(
                "Self Test",
                self._create_self_test_headers
            )

            # Get the next available row
            row_num = self.self_test_sheet.max_row + 1

            # Write data with timestamp
            self.self_test_sheet.cell(
                row=row_num, column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            self.self_test_sheet.cell(row=row_num, column=2, value=unit_identifier)

            # Test Result with color coding
            result_cell = self.self_test_sheet.cell(
                row=row_num, column=3,
                value="PASS" if test_passed else "FAIL"
            )
            result_cell.font = Font(bold=True)
            result_cell.fill = PatternFill(
                start_color="00AA00" if test_passed else "FF0000",
                end_color="00AA00" if test_passed else "FF0000",
                fill_type="solid"
            )
            result_cell.alignment = Alignment(horizontal='center')

            self.self_test_sheet.cell(row=row_num, column=4, value=test_details)
            self.self_test_sheet.cell(row=row_num, column=5, value=notes)

            # Save the workbook
            self.workbook.save(self.file_path)
            self.logger1.info(f"Logged self test result to {self.file_path}",
                              extra={'func_name': 'Log_self_test'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging self test result: {e}", exc_info=True,
                               extra={'func_name': 'Log_self_test'})
            return False

    @log_function
    def log_resistance_measurement(self, measurement_data, sheet_name):
        """Log resistance measurement to combined sheet"""
        try:
            # Ensure sheet exists
            self.resistance_sheet = self._ensure_sheet_exists(
                sheet_name,
                self._create_resistance_headers
            )

            # Write measurement data
            row_num = self.resistance_sheet.max_row + 1
            self.resistance_sheet.cell(row=row_num, column=1, value=measurement_data['timestamp'])
            self.resistance_sheet.cell(row=row_num, column=2, value=measurement_data['zone_title'])
            self.resistance_sheet.cell(row=row_num, column=3, value=measurement_data['setpoint'])
            self.resistance_sheet.cell(row=row_num, column=4, value=measurement_data['resistance'])

            # Status with color coding
            status_cell = self.resistance_sheet.cell(
                row=row_num, column=5,
                value=measurement_data['status']
            )
            if measurement_data['status'] == "PASS":
                status_cell.fill = PatternFill(
                    start_color="00AA00",  # Green
                    fill_type="solid"
                )
            else:
                status_cell.fill = PatternFill(
                    start_color="FF0000",  # Red
                    fill_type="solid"
                )

            self.resistance_sheet.cell(row=row_num, column=6, value=measurement_data['table_row'])

            # Save workbook
            self.workbook.save(self.file_path)
            self.logger1.info(
                f"Logged resistance data to combined worksheet",
                extra={'func_name': 'log_resistance_measurement'}
            )
            return True
        except Exception as e:
            self.logger1.error(
                f"Error logging resistance data: {str(e)}",
                exc_info=True,
                extra={'func_name': 'log_resistance_measurement'}
            )
            return False


    @log_function
    def log_summary(self, metadata=None, teststep_data=None, step_data=None, update_existing=True):
        """Log data to the Summary sheet with option to update existing entries"""
        try:
            # Ensure sheet exists
            summary_sheet = self._ensure_sheet_exists(
                "Summary",
                self._create_summary_headers
            )

            # Update metadata if provided
            if metadata:
                metadata_mapping = {
                    'eid': 1,
                    'serial_number': 2,
                    'model_number': 3,
                    'version': 4,
                    'tester_name': 5,
                    'comment': 6,
                    'start_time': 7,
                    'end_time': 8,
                    'overall_result': 9,
                    'test_fixture_sn': 10,
                    'vna_sn': 11,
                    'ecal_sn': 12
                }

                for key, value in metadata.items():
                    if key.lower() in metadata_mapping:
                        row_num = metadata_mapping[key.lower()]
                        summary_sheet.cell(row=row_num, column=2, value=value)

                        if key.lower() == 'overall_result':
                            result_cell = summary_sheet.cell(row=row_num, column=2)
                            result_cell.font = Font(bold=True)
                            result_cell.alignment = Alignment(horizontal='center')
                            if str(value).upper() == "PASS":
                                result_cell.fill = PatternFill(
                                    start_color="00AA00", end_color="00AA00", fill_type="solid")
                                result_cell.font = Font(color="FFFFFF", bold=True)
                            elif str(value).upper() == "FAIL":
                                result_cell.fill = PatternFill(
                                    start_color="FF0000", end_color="FF0000", fill_type="solid")
                                result_cell.font = Font(color="FFFFFF", bold=True)

            # Update TESTSTEP data directly if provided
            if teststep_data:
                teststep_name = teststep_data.get('teststep', '')
                teststep_status = teststep_data.get('status', '')

                if teststep_name:
                    # Find existing teststep in TESTSTEP section (rows 15-20, column 1)
                    teststep_updated = False
                    for row in range(15, 21):  # TESTSTEP section rows
                        existing_teststep = summary_sheet.cell(row=row, column=1).value
                        if existing_teststep == teststep_name:
                            # Update existing teststep
                            status_cell = summary_sheet.cell(row=row, column=2, value=teststep_status)
                            self._apply_status_formatting(status_cell, teststep_status)
                            teststep_updated = True
                            break

                    # If not found and we have space, add to first empty row
                    if not teststep_updated:
                        for row in range(15, 21):
                            existing_teststep = summary_sheet.cell(row=row, column=1).value
                            if not existing_teststep or existing_teststep == "":
                                # Add new teststep
                                summary_sheet.cell(row=row, column=1, value=teststep_name)
                                status_cell = summary_sheet.cell(row=row, column=2, value=teststep_status)
                                self._apply_status_formatting(status_cell, teststep_status)
                                teststep_updated = True
                                break

            # Log STEP data if provided
            if step_data:
                testpoints = step_data.get('testpoints', '')

                # Find existing row if update_existing is True - search in TestPoints column (column G, index 7)
                step_row = None
                if update_existing and testpoints:
                    # Search from row 23 onwards for matching testpoints
                    for row in range(23, summary_sheet.max_row + 1):
                        cell_value = summary_sheet.cell(row=row, column=7).value  # Column G (7) is TestPoints
                        if cell_value == testpoints:
                            step_row = row
                            break

                # If not found or not updating, use next available row
                if step_row is None:
                    step_row = 23
                    # Find first empty row in the STEP section
                    while (summary_sheet.cell(row=step_row, column=7).value is not None and
                           summary_sheet.cell(row=step_row, column=7).value != ""):
                        step_row += 1

                    # For new rows, write step data including the step column
                    summary_sheet.cell(row=step_row, column=1, value=step_data.get('step', ''))  # Column A: Step
                # For existing rows, DO NOT update the Step column (Column A) - keep it as is

                # Update columns B to H (Unit to Status) - preserving Step column (A)
                summary_sheet.cell(row=step_row, column=2, value=step_data.get('unit', ''))  # Column B: Unit
                summary_sheet.cell(row=step_row, column=3, value=step_data.get('low_limit', ''))  # Column C: Low_Limit
                summary_sheet.cell(row=step_row, column=4, value=step_data.get('measure', ''))  # Column D: Measure
                summary_sheet.cell(row=step_row, column=5,
                                   value=step_data.get('high_limit', ''))  # Column E: High_Limit
                summary_sheet.cell(row=step_row, column=6, value=step_data.get('teststep', ''))  # Column F: TestStep
                summary_sheet.cell(row=step_row, column=7, value=testpoints)  # Column G: TestPoints



                # Write STATUS with color coding in column H (index 8)
                status = step_data.get('status', '')
                status_cell = summary_sheet.cell(row=step_row, column=8, value=status)  # Column H: Status
                self._apply_status_formatting(status_cell, status)

            # If step_data was provided, update TESTSTEP section automatically based on STEP section data
            # BUT only update teststeps that are managed by STEP data (not manually set ones)
            if step_data:
                self._update_teststep_from_step_data_preserve_manual(summary_sheet)

            # AUTOMATICALLY UPDATE OVERALL RESULT BASED ON TESTSTEP STATUS (ROWS 15-20, COLUMN B)
            self._update_overall_result_based_on_teststep_status(summary_sheet)

            # Save the workbook
            self.workbook.save(self.file_path)
            self.logger1.info("Logged summary data successfully",
                              extra={'func_name': 'log_summary'})
            return True

        except Exception as e:
            self.logger1.error(f"Error logging summary data: {e}", exc_info=True,
                               extra={'func_name': 'log_summary'})
            return False

    @log_function
    def _update_teststep_from_step_data_preserve_manual(self, summary_sheet):
        """Update TESTSTEP section based on STEP section data but preserve manually set teststeps"""
        try:
            # Get current manually set teststeps (rows 15-20)
            manual_teststeps = {}
            for row in range(15, 21):
                teststep_name = summary_sheet.cell(row=row, column=1).value
                if teststep_name:
                    manual_teststeps[teststep_name] = {
                        'row': row,
                        'status': summary_sheet.cell(row=row, column=2).value
                    }

            # Dictionary to store teststep statuses from STEP data
            step_teststep_status_map = {}

            # Collect all teststeps and their statuses from STEP section (column F and H)
            for row in range(23, summary_sheet.max_row + 1):
                teststep_cell = summary_sheet.cell(row=row, column=6)  # Column F: TestStep
                status_cell = summary_sheet.cell(row=row, column=8)  # Column H: Status

                teststep_name = teststep_cell.value
                status_value = str(status_cell.value).upper().strip() if status_cell.value else ""

                # Only process teststeps that are NOT manually managed
                if teststep_name and teststep_name not in manual_teststeps:
                    if teststep_name not in step_teststep_status_map:
                        step_teststep_status_map[teststep_name] = "PASS"  # Start with PASS assumption

                    # If any step in a teststep fails, the entire teststep fails
                    if status_value == "FAIL":
                        step_teststep_status_map[teststep_name] = "FAIL"

            # Update TESTSTEP section (rows 15-20) - only for step-managed teststeps
            current_teststep_row = 15

            # First, keep existing manual teststeps
            for row in range(15, 21):
                teststep_name = summary_sheet.cell(row=row, column=1).value
                if teststep_name and teststep_name in manual_teststeps:
                    # Keep manual teststep as is
                    current_teststep_row += 1

            # Then add step-managed teststeps to remaining rows
            for teststep_name, status in step_teststep_status_map.items():
                if current_teststep_row > 20:  # Don't exceed the TESTSTEP section
                    break

                # Write teststep name
                summary_sheet.cell(row=current_teststep_row, column=1, value=teststep_name)

                # Write status with color coding
                status_cell = summary_sheet.cell(row=current_teststep_row, column=2, value=status)
                self._apply_status_formatting(status_cell, status)

                current_teststep_row += 1

            # Clear any remaining rows in TESTSTEP section
            for row in range(current_teststep_row, 21):
                # Only clear if not a manual teststep
                existing_teststep = summary_sheet.cell(row=row, column=1).value
                if existing_teststep not in manual_teststeps:
                    summary_sheet.cell(row=row, column=1, value="")
                    summary_sheet.cell(row=row, column=2, value="")

            self.logger1.info(
                f"Updated TESTSTEP section - Manual: {len(manual_teststeps)}, Step-managed: {len(step_teststep_status_map)}",
                extra={'func_name': '_update_teststep_from_step_data_preserve_manual'})

        except Exception as e:
            self.logger1.error(f"Error updating TESTSTEP from STEP data: {e}", exc_info=True,
                               extra={'func_name': '_update_teststep_from_step_data_preserve_manual'})

    @log_function
    def _apply_status_formatting(self, cell, status):
        """Apply consistent status formatting to a cell"""
        status = str(status).upper().strip() if status else ""
        cell.font = Font(bold=True)

        if status == "PASS":
            cell.fill = PatternFill(start_color="00AA00", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        elif status == "FAIL":
            cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

    @log_function
    def _update_overall_result_based_on_teststep_status(self, summary_sheet):
        """Update overall result based only on TESTSTEP status (rows 15-20, column B)"""
        try:
            overall_result = "PASS"  # Start with PASS assumption

            # Check TESTSTEP status from rows 15-20, column B
            for row in range(15, 21):  # TESTSTEP section rows 15-20
                status_cell = summary_sheet.cell(row=row, column=2)  # Column B: Status
                status_value = str(status_cell.value).upper().strip() if status_cell.value else ""

                # If any teststep status is FAIL or empty, overall result becomes FAIL
                if status_value == "FAIL" or status_value == "":
                    overall_result = "FAIL"
                    break  # No need to check further if we found one FAIL

            # Update overall result in row 9, column 2
            result_cell = summary_sheet.cell(row=9, column=2, value=overall_result)
            result_cell.font = Font(bold=True)
            result_cell.alignment = Alignment(horizontal='center')
            self._apply_status_formatting(result_cell, overall_result)

            self.logger1.info(f"Updated overall result to: {overall_result} (based on TESTSTEP status)",
                              extra={'func_name': '_update_overall_result_based_on_teststep_status'})

        except Exception as e:
            self.logger1.error(f"Error updating overall result from teststep status: {e}", exc_info=True,
                               extra={'func_name': '_update_overall_result_based_on_teststep_status'})

    @log_function
    def log_Imp_measurement(self, measurement_data, sheet_name):
        """Log impedance measurement to combined sheet"""
        try:
            # Ensure sheet exists
            self.Imp_sheet = self._ensure_sheet_exists(
                sheet_name,
                self._create_impedance_headers
            )

            row_num = self.Imp_sheet.max_row + 1
            self.Imp_sheet.cell(row=row_num, column=1, value=measurement_data['timestamp'])
            self.Imp_sheet.cell(row=row_num, column=2, value=measurement_data['zone_title'])
            self.Imp_sheet.cell(row=row_num, column=3, value=measurement_data['Frequency'])
            self.Imp_sheet.cell(row=row_num, column=4, value=measurement_data['setpoint'])
            self.Imp_sheet.cell(row=row_num, column=5, value=measurement_data['Real'])
            self.Imp_sheet.cell(row=row_num, column=6, value=measurement_data['Imag'])
            self.Imp_sheet.cell(row=row_num, column=7, value=measurement_data['Z'])
            self.Imp_sheet.cell(row=row_num, column=8, value=measurement_data['status'])

            # Status with color coding
            status_cell = self.Imp_sheet.cell(
                row=row_num, column=8,
                value=measurement_data['status']
            )
            if measurement_data['status'] == "PASS":
                status_cell.fill = PatternFill(
                    start_color="00AA00",  # Green
                    fill_type="solid"
                )
            else:
                status_cell.fill = PatternFill(
                    start_color="FF0000",  # Red
                    fill_type="solid"
                )

            # Save workbook
            self.workbook.save(self.file_path)
            self.logger1.info(
                f"Logged impedance data to combined worksheet",
                extra={'func_name': 'log_impedance_measurement'}
            )
            return True
        except Exception as e:
            self.logger1.error(
                f"Error logging impedance data: {str(e)}",
                exc_info=True,
                extra={'func_name': 'log_impedance_measurement'}
            )
            return False



    @log_function
    def log_BNC_measurement(self, test_zone, test_details, test_passed):
        """Log BNC measurement to combined sheet"""
        try:
            # Ensure sheet exists
            self.BNC_sheet = self._ensure_sheet_exists(
                "BNC Port Verification",
                self._create_BNC_headers
            )

            # Get the next available row
            row_num = self.BNC_sheet.max_row + 1

            # Write data with timestamp
            self.BNC_sheet.cell(
                row=row_num, column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            self.BNC_sheet.cell(row=row_num, column=2, value=test_zone)
            self.BNC_sheet.cell(row=row_num, column=3, value=test_details)
            # Test Result with color coding
            result_cell = self.BNC_sheet.cell(
                row=row_num, column=4,
                value="PASS" if test_passed else "FAIL"
            )
            result_cell.font = Font(bold=True)
            result_cell.fill = PatternFill(
                start_color="00AA00" if test_passed else "FF0000",
                end_color="00AA00" if test_passed else "FF0000",
                fill_type="solid"
            )
            result_cell.alignment = Alignment(horizontal='center')

            # Save the workbook
            self.workbook.save(self.file_path)
            self.logger1.info(f"Logged BNC test result to {self.file_path}",
                              extra={'func_name': 'Log_BNC_test'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging BNC test result: {e}", exc_info=True,
                               extra={'func_name': 'Log_BNC_test'})
            return False


# Initialize the Excel logger
excel_logger = ExcelLogger()
